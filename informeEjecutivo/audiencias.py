from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from os import listdir
import datetime
from openpyxl.styles import PatternFill, Alignment

def audiencias(path, template):
    """
    Funcion para condensar los archivos de audiencias en uno solo. Lee todos los archivos de una carpeta seleccionada
    y agrega en una hoja todos los registros de audiencias ordenados por fecha
    :param path: El directorio en donde se encuentran los archivos
    :param template: El archivo usado como formato para llenar con la informacion de los archivos individuales
    :return:
    """
    filenames = listdir(path)
    space_counter = 0
    date_dict = {}
    for file in filenames:
        wb = load_workbook(f'{path}{file}')
        sheet = wb["Audiencias Acumuladas"]
        hidden_rows = []
        for row_num, row_dim in sheet.row_dimensions.items():
            if row_dim.hidden:
                hidden_rows.append(row_num)
        for row in sheet.iter_rows(min_row=2):
            if row[1].row not in hidden_rows:
                row_list = [cell.value for cell in row]
                if isinstance(row_list[7], datetime.datetime) or isinstance(row_list[7], datetime.date):
                    row_list[7] = row_list[7].strftime("%I:%M")
                if isinstance(row_list[6], datetime.datetime) or isinstance(row_list[6], datetime.date):
                    row_list[6] = row_list[6].strftime("%d/%m/%Y")
                if isinstance(row_list[5], datetime.datetime) or isinstance(row_list[5], datetime.date):
                    row_list[5] = row_list[5].strftime("%d/%m/%Y")
                if row_list[1] is None:
                    space_counter += 1
                    if space_counter >= 10:
                        break
                elif row_list[6] not in date_dict:
                    date_dict.update({row_list[6]: []})
                    date_dict[row_list[6]].append(row_list)
                else:
                    date_dict[row_list[6]].append(row_list)
                # print(date_dict)
    # pprint.pprint(date_dict)

    summ_wb = load_workbook(f'{template}')
    summ_sheet = summ_wb["Audiencias Acumuladas"]
    current_row = 2
    for key, value in date_dict.items():
        summ_sheet.merge_cells(start_row=current_row, end_row=current_row, start_column=1, end_column=12)
        summ_sheet[f'A{current_row}'].value = key
        summ_sheet[f'A{current_row}'].fill = PatternFill(start_color="9dc3e6", end_color="9dc3e6", fill_type="solid")
        summ_sheet[f'A{current_row}'].alignment = Alignment(horizontal="center")
        current_row += 1
        for val in value:
            for index2, field in enumerate(val):
                summ_sheet[f'{get_column_letter(index2 + 1)}{current_row}'].value = field
            current_row += 1

    summ_wb.save(f'{path}{datetime.date.today().strftime("%Y%m%d")} Audiencias.xlsx')