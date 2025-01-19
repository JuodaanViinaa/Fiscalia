from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from os import listdir
import datetime
from openpyxl.styles import PatternFill, Alignment

def ordenes(path, template):
    """
    Funcion para condensar los archivos de ordenes en uno solo. Lee todos los archivos de una carpeta seleccionada
    y agrega en una hoja todos los registros de ordenes por fiscalia
    :param path: El directorio en donde se encuentran los archivos
    :param template: El archivo usado como formato para llenar con la informacion de los archivos individuales
    :return:
    """
    filenames = listdir(path)
    space_counter = 0
    master_list = []
    for file in filenames:
        wb = load_workbook(f'{path}{file}')
        sheet = wb.active
        for row in sheet.iter_rows(min_row=10):
            row_list = [cell.value for cell in row]
            if row_list[1] is None:
                space_counter += 1
                if space_counter >= 10:
                    break
            # elif "novedad" not in str(row_list[0]).lower and "novedad" not in str(row_list[1]).lower:
            else:
                master_list.append(row_list)

    summ_wb = load_workbook(f'{template}')
    summ_sheet = summ_wb.active
    current_row = 10
    for sublist in master_list:
        for index, item in enumerate(sublist):
            summ_sheet[f'{get_column_letter(index + 1)}{current_row}'].value = item
        current_row += 1
    summ_wb.save(f'{path}{datetime.date.today().strftime("%Y%m%d")} Ordenes.xlsx')