from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import datetime

def informe_ejecutivo(dir_path, template):
    """
    Funcion para realizar el informe ejecutivo diario.
    Lee los archivos de la carpeta provista y determina su pertenencia a cada fiscalia con base en dos criterios:
    primero, el nombre del archivo, lo que es rapido y eficiente; si queda algun archivo sin asignar se utiliza el
    criterio de contenido.
    Los resumenes de cada fiscalia se copian hacia el archivo final, y despues todos los archivos son renombrados
    para ser consistentes y faciles de identificar.
    :param dir_path: El directorio en donde se encuentran todos los archivos sin clasificar antes del informe
    :param template: El archivo vacio usado como formato para llenar con los datos de cada fiscalia
    :return:
    """
    filenames = sorted(os.listdir(dir_path))
    # El diccionario almacenará el nombre de cada uno de los archivos
    # y lo relacionará con su fiscalía correspondiente.
    file_dict = {
        'FIDS': None,
        'FIDCANNA': None,
        'FEIDF': None,
        'FIDMTP': None,
        'FIDVF': None,
        'FIDAGAP': None,
        'FJPA': None,
        'CJM': None,
         }
    fisc_nombres = {
        'FIDS': "Fiscalía Central de Investigación para la Atención de Delitos  Sexuales.  ",
        'FIDCANNA': "Fiscalía de Investigación de Delitos Cometidos en Agravio de Niños, Niñas y Adolescentes.",
        'FEIDF': "Fiscalía  Especializada para la Investigación del Delito de Feminicidio.    ",
        'FIDMTP': "Fiscalía Central de Investigación para la Atención del Delito de Trata de Personas ",
        'FIDVF': "Fiscalía de Investigacion del delito de Violencia Familiar",
        'FIDAGAP': "Fiscalía de Investigación de Delitos Cometidos en Agravio de Grupos de Atención Prioritaria",
        'FJPA': "Fiscalía de Justicia Para Adolescentes",
    }
    # Se busca dentro de los archivos de la carpeta por coincidencias en el nombre.
    # Las coincidencias se agregan al diccionario de arriba y a una lista de
    # archivos presentes.
    archivos_presentes = []
    for file in filenames:
        wb = load_workbook(f'{dir_path}{file}')
        if "FIDCANNA" in file:
            file_dict["FIDCANNA"] = file
            archivos_presentes.append(file)
        elif "FIDMTP" in file:
            file_dict["FIDMTP"] = file
            archivos_presentes.append(file)
        elif " VF " in file or "FIDVF" in file:
            file_dict["FIDVF"] = file
            archivos_presentes.append(file)
        elif "FJPA" in file:
            file_dict["FJPA"] = file
            archivos_presentes.append(file)
        elif ("FIDAGAP" in file) or ("Prioritaria" in file):
            file_dict["FIDAGAP"] = file
            archivos_presentes.append(file)
        elif "CJM" in file:
            file_dict["CJM"] = file
            archivos_presentes.append(file)
        elif "FDS" in file or "FIDS" in file:
            file_dict["FIDS"] = file
            archivos_presentes.append(file)
        elif "FEMI" in file:
            file_dict["FEIDF"] = file
            archivos_presentes.append(file)

    # Los archivos que no se pudieron asociar a una fiscalía se almacenan en
    # una lista que será revisada de nuevo para buscar ahora dentro de cada
    # uno y encontrar coincidencias.
    # Se revisa el interior de cada archivo en una celda en la cual haya algo
    # único que lo identifique. Si se encuentra, el archivo se agrega al diccionario
    # y a la lista de archivos presentes.
    archivos_sobrantes = [file for file in filenames if file not in archivos_presentes]
    fiscalias_faltantes = [key for key, value in file_dict.items() if value is None]
    if len(fiscalias_faltantes) > 0:
            for file in archivos_sobrantes:
                wb = load_workbook(f'{dir_path}{file}')
                try:
                    if "AS Y ADOLESCENTES" in str(wb["Concentrado"]["A1"].value):
                        file_dict["FIDCANNA"] = file
                    elif "TRATA DE PERSONAS" in str(wb["Concentrado"]["A1"].value):
                        file_dict["FIDMTP"] = file
                    elif "AGENCIA 75" in str(wb["Acciones"]["B1"].value):
                        file_dict["FIDVF"] = file
                    elif "FJPA" in str(wb["Acciones"]["B1"].value):
                        file_dict["FJPA"] = file
                    elif "PRIORITARIA" in str(wb["Concentrado"]["A1"].value):
                        file_dict["FIDAGAP"] = file
                    elif "FDS-1" in wb["Acciones"]["B1"].value:
                        file_dict["FIDS"] = file
                    elif "FEMINICIDIO" in str(wb["Concentrado"]["A1"].value):
                        file_dict["FEIDF"] = file
                except:
                    if "CONCEPTO" in str(wb.active["B3"].value):
                        file_dict["CJM"] = file

    # Se crea una lista con las fiscalias que no enviaron reporte y se imprime un mensaje
    fiscalias_faltantes = [key for key, value in file_dict.items() if value is None]
    if len(fiscalias_faltantes) > 0:
        print(f'Estas fiscalías no enviaron su reporte o su archivo no se pudo encontrar: {fiscalias_faltantes}')
    else:
        print("Todas las fiscalías enviaron su reporte. :)")

    # Se compila una lista nueva con las fiscalias no presentes en la lista de faltantes
    # Se cicla a través de esa lista en el cuerpo principal del codigo
    file_list = [fisc for fisc in ['FIDS', 'FIDCANNA', 'FEIDF', 'FIDMTP', 'FIDVF', 'FIDAGAP', 'FJPA'] if fisc not in fiscalias_faltantes]

    summ_wb = load_workbook(f'{template}')
    # Se separan las celdas unidas de la hoja de Acciones
    summ_wb["Acciones"].unmerge_cells(start_row=1, start_column=1, end_row=3, end_column=9)
    summ_wb["Acciones"].unmerge_cells(start_row=4, start_column=1, end_row=4, end_column=9)
    for fiscalia in file_list:
        wb = load_workbook(f'{dir_path}{file_dict[fiscalia]}', data_only=True)
        print(f'Trabajando en {fiscalia}')

        # Acciones
        column_fisc = None
        column_summ = None
        for col in wb["Acciones"].iter_cols(1):
            for cell in col:
                if cell.value in ["TOTAL", "FJPA", "FIDMTP", "FIDAGAP", "FEIDF"]:
                    column_fisc = cell.column
                    break
                break
        for col in summ_wb["Acciones"].iter_cols(min_row=6, max_row=6):
            for cell in col:
                if cell.value == fiscalia:
                    column_summ = cell.column
                    break
                break
        for index, cell in enumerate(summ_wb["Acciones"]["A"]):
            for index2, cell2 in enumerate(wb["Acciones"]["A"]):
                if cell.value == cell2.value and cell.value != "ACCIONES":
                    summ_wb["Acciones"][f'{get_column_letter(column_summ)}{index + 1}'].value = wb["Acciones"][f'{get_column_letter(column_fisc)}{index2 + 1}'].value
                    break
        print("    Acciones terminado")

        # Medidas de protección
        column_fisc = None
        column_summ = None
        for col in wb["Medidas de Protección"].iter_cols(1):
            for cell in col:
                if cell.value in ["TOTAL", "FJPA", "FIDMTP", "FIDMT", "FIDAGAP", "FEIDF"]:
                    column_fisc = cell.column
                    break
                break
        for row in summ_wb["Medidas de Protección"].iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value == fiscalia:
                    column_summ = cell.column
                    break
                # break  # ¿Es necesario este break? Ponerlo rompe el código, pero antes no lo hacía
        for index, cell in enumerate(summ_wb["Medidas de Protección"]["A"]):
            for index2, cell2 in enumerate(wb["Medidas de Protección"]["A"]):
                if cell.value == cell2.value and cell.value != "ACCIONES":
                    summ_wb["Medidas de Protección"][f'{get_column_letter(column_summ)}{index + 1}'].value = wb["Medidas de Protección"][f'{get_column_letter(column_fisc)}{index2 + 1}'].value
                    break
        print("    Medidas de Protección terminado")

        # Concentrado
        row_fisc = None
        row_summ = None
        for row in wb["Concentrado"].iter_rows(1):
            for cell in row:
                if cell.value in ["TOTAL"]:
                    row_fisc = cell.row
                    break
                break
        for row in summ_wb["Concentrado"].iter_rows(min_row=1, max_row=9):
            for cell in row:
                if cell.value == fisc_nombres[fiscalia]:
                    row_summ = cell.row
                    break
                break
        wb["Concentrado"].unmerge_cells(start_row=row_fisc, start_column=2, end_row=row_fisc, end_column=3)
        wb["Concentrado"].unmerge_cells(start_row=row_fisc, start_column=4, end_row=row_fisc, end_column=5)
        summ_wb["Concentrado"][f'B{row_summ}'].value = wb["Concentrado"][f'B{row_fisc}'].value
        summ_wb["Concentrado"][f'C{row_summ}'].value = wb["Concentrado"][f'D{row_fisc}'].value
        summ_wb["Concentrado"]["B9"].value = "=SUM(B2:B8)"
        summ_wb["Concentrado"]["C9"].value = "=SUM(C2:C8)"
        print("    Concentrado terminado")

        # Libertades
        current_row = 2
        for row in wb["Libertades"].iter_rows(min_row=2):
            for row2 in summ_wb["Libertades"].iter_rows(max_row=summ_wb["Libertades"].max_row + 1):
                if row2[0].value is None:
                    current_row = row2[0].row
                    break
            row_values = [cell.value for cell in row]
            for index, value in enumerate(row_values):
                if value is not None:
                # if value is not None and ("NOVEDAD" not in value.upper()):
                    summ_wb["Libertades"][current_row][index].value = value
        print(f'    Libertades terminado')

        # Audiencias
        divisiones = [
            "AUDIENCIAS ATENDIDAS POR LAS FISCALIAS DE LA CGIDGAV – Asesor Estrategias procesales",
            "AUDIENCIAS ATENDIDAS POR LAS FISCALIAS DE LA CGIDGAV – Asesor Estrategias procesales",
            "AUDIENCIAS ATENDIDAS POR LAS FISCALIAS DE LA CGIDGAV – Asesor de Ejecución"
        ]
        column_summ = None
        for col in summ_wb["Audiencias"].iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value == fiscalia:
                    column_summ = cell.column
                    break
                break
        for cell in wb["Audiencias"]["A"]:
            if cell.value == "Audiencia de prórroga del plazo de cierre de complenetaria ":
                cell.value = "Audiencia de prórroga del plazo de cierre de complementaria "
            elif cell.value == "Audiencia Inicial sin detenido":
                cell.value = "Audiencia inicial sin detenido "
            elif cell.value == "Audiencia de Actos de InvestIgación":
                cell.value = "Audiencia de solicitud de actos de investigación"
        for index, cell in enumerate(summ_wb["Audiencias"]["A"]):
            for index2, cell2 in enumerate(wb["Audiencias"]["A"]):
                if cell.value == cell2.value and cell.row > 1 and cell2.value not in divisiones:
                    summ_wb["Audiencias"][f'{get_column_letter(column_summ)}{index + 1}'].value = wb["Audiencias"][f'B{index2 + 1}'].value
                    break
        print("    Audiencias terminado")
        summ_wb.save(f'{dir_path}{datetime.datetime.now().strftime("%Y%m%d")} INFORME EJECUTIVO DIARIO CGIDGAV.xlsx')

    # CJM
    cjm_wb = load_workbook(f'{dir_path}{file_dict["CJM"]}', data_only=True)
    values1 = []
    for row in range(4, 15):
        for col in ["C", "D", "E", "F", "G"]:
            values1.append(cjm_wb.active[f'{col}{row}'].value)
    for row in range(4, 15):
        for col in ["B", "C", "D", "E", "F"]:
            summ_wb["CJM"][f'{col}{row}'].value = values1.pop(0)
    values2 = []
    for row in range(19, 29):
        for col in ["C", "D", "E", "F", "G"]:
            values2.append(cjm_wb.active[f'{col}{row}'].value)
    for row in range(19, 29):
        for col in ["B", "C", "D", "E", "F"]:
            summ_wb["CJM"][f'{col}{row}'].value = values2.pop(0)
    print("CJM Terminado")

    # Los archivos se renombran con un nombre más fácil e informativo
    for name, file in file_dict.items():
        if file is not None:
            os.rename(f'{dir_path}{file}', f'{dir_path}{datetime.date.today().strftime("%Y%m%d")} {name} Informe Ejecutivo.xlsx')

    summ_wb.save(f'{dir_path}{datetime.datetime.now().strftime("%Y%m%d")} INFORME EJECUTIVO DIARIO CGIDGAV.xlsx')